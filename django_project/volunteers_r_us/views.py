from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from types import SimpleNamespace
from django.contrib.auth import get_user_model
from .models import Skill, Event
from datetime import datetime, date
from datetime import date
from django.shortcuts import render
from .models import Assignment
from notify.models import Notification
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout, get_user_model
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from io import BytesIO
from datetime import datetime as _dt
from .models import Profile
from .forms import ProfileForm, EventForm, STATE_CHOICES, SKILL_CHOICES
import logging
from django.conf import settings

#from .models import VolunteerParticipation as VP
from django.utils.timezone import now
from django.contrib.admin.views.decorators import staff_member_required


User = get_user_model()

def home(request):
    return render(request, 'home.html')

def login_view(request):
    # Renders form on GET; authenticates on POST
    ctx = {}
    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        # Basic field validations (Assignment requirement)
        errors = []
        if not email:
            errors.append("Email is required.")
        if not password:
            errors.append("Password is required.")
        if errors:
            ctx["info"] = " ".join(errors)
            return render(request, "login.html", ctx)

        user = authenticate(request, username=email, password=password)
        if user is not None:
            auth_login(request, user)
            return redirect(request.GET.get("next") or "home")
        ctx["info"] = "Invalid email or password."
    return render(request, "login.html", ctx)

def register_view(request):
    # Renders form on GET; creates user on POST
    ctx = {}
    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")
        confirm = request.POST.get("confirm", "")

        # Server-side validations (required fields, types/length)
        errors = []
        if not email:
            errors.append("Email is required.")
        if not password:
            errors.append("Password is required.")
        if password and len(password) < 8:
            errors.append("Password must be at least 8 characters.")
        if password != confirm:
            errors.append("Passwords must match.")
        if User.objects.filter(email__iexact=email).exists():
            errors.append("An account with that email already exists.")

        # Run Django’s password validators too
        if password:
            try:
                validate_password(password)
            except ValidationError as ve:
                errors.extend(ve.messages)

        if errors:
            ctx["info"] = " ".join(errors)
            return render(request, "register.html", ctx)

        # Create user and redirect to login
        User.objects.create_user(email=email, password=password)
        ctx["info"] = "Account created. Please sign in."
        return redirect("login")

    return render(request, "register.html", ctx)

def logout_view(request):
    auth_logout(request)
    return redirect("home")

#@login_required
# volunteers_r_us/views.py
from django.shortcuts import render

def _parse_date(s: str):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _skills_list(event):
    """Return list of skill names from Event.required_skills (ManyToManyField)."""
    if not event:
        return []
    return [str(skill).strip() for skill in event.required_skills.all()]


def _csv_from_list(values):
    return ", ".join(v for v in values if v)

from io import BytesIO  

from io import BytesIO

def _pdf_table_response(columns, rows, filename_base="report", filename_prefix=None):
    """
    Generic PDF table helper used by:
      - volunteer_history (10 columns)
      - event_form (6 columns)

    It:
      * Fits the table to the page width
      * Gives more room to text-heavy columns (Description, Required Skills, Capacity)
      * Wraps long text inside cells so headers/values stay inside their cells
    """
    if filename_prefix is None:
        filename_prefix = filename_base

    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = BytesIO()
    page_size = landscape(letter)
    page_width, page_height = page_size

    # Margins and usable width
    left_margin = right_margin = 30
    top_margin = bottom_margin = 30
    usable_width = page_width - (left_margin + right_margin)

    # Build raw data: header row + data rows
    data = [columns] + rows
    num_cols = len(columns)

    # --------- Column width weights (relative, not pixels) ----------
    if num_cols == 10:
        # volunteer_history: [Volunteer, Event, Description, Location, Required Skills,
        #                     Urgency, Event Date, Capacity, Languages, Status]
        weights = [
            1.1,  # Volunteer
            1.0,  # Event Name
            1.5,  # Description
            0.8,  # Location
            2.3,  # Required Skills  (wide)
            0.7,  # Urgency
            1.0,  # Event Date
            2.0,  # Capacity (current / total)
            0.8,  # Languages
            0.6,  # Status
        ]
    elif num_cols == 6:
        # events: [Event Name, Description, Location, Required Skills, Urgency, Event Date]
        weights = [
            1.3,  # Event Name
            1.7,  # Description
            1.0,  # Location
            2.4,  # Required Skills (wide)
            0.7,  # Urgency
            0.9,  # Event Date
        ]
    else:
        # Fallback: equal widths
        weights = [1.0] * num_cols

    total_weight = sum(weights)
    col_widths = [usable_width * (w / total_weight) for w in weights]

    # --------- Convert all cells to Paragraphs so they wrap ---------
    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]
    body_style.fontSize = 8
    body_style.leading = 10
    body_style.wordWrap = "CJK"

    header_style = styles["BodyText"]
    header_style.fontSize = 8
    header_style.leading = 10
    header_style.wordWrap = "CJK"
    header_style.fontName = "Helvetica-Bold"

    para_data = []
    for row_idx, row in enumerate(data):
        new_row = []
        for cell in row:
            text = "" if cell is None else str(cell)
            if row_idx == 0:
                new_row.append(Paragraph(text, header_style))
            else:
                new_row.append(Paragraph(text, body_style))
        para_data.append(new_row)

    # --------- Create document & table ----------
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    table = Table(para_data, colWidths=col_widths, repeatRows=1)

    # --------- Styling (grid, alignment, etc.) ----------
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    doc.build([table])

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename_prefix}.pdf"'
    return response


@login_required
@staff_member_required
def volunteer_history(request):
    if request.GET.get("reset") == "1":
        return redirect(request.path)

    # ---------- filters from query string ----------
    volunteer = (request.GET.get("volunteer") or "").strip()  # volunteer email
    status    = (request.GET.get("status") or "").strip()
    from_str  = (request.GET.get("from_date") or "").strip()

    d_from = _parse_date(from_str)

    # Base queryset: assignments with volunteer and event joined
    qs = Assignment.objects.select_related("volunteer", "event").all()

    if volunteer:
        qs = qs.filter(volunteer__email=volunteer)
    if status:
        qs = qs.filter(status=status)
    if d_from:
        qs = qs.filter(event__event_date=d_from)

    qs = qs.order_by("event__event_date", "volunteer__email")

    # ------------------------------- export branch -------------------------------
    if request.GET.get("export") == "1":
        headers = [
            "Volunteer", "Event Name", "Description", "Location",
            "Required Skills", "Urgency", "Event Date",
            "Capacity (current / total)", "Languages", "Status",
        ]
        filename = "volunteer_history"
        if volunteer:
            filename += f"_{volunteer}"
        if status:
            filename += f"_{status}"
        stamp = now().strftime("%Y%m%d-%H%M%S")

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Volunteer History"
            ws.append(headers)

            # header style
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for rec in qs:
                evt = rec.event
                user = rec.volunteer

                # Full name if available, otherwise email
                if hasattr(user, "profile") and getattr(user.profile, "full_name", "").strip():
                    volunteer_name = user.profile.full_name.strip()
                else:
                    volunteer_name = getattr(user, "email", str(user))

                skills_csv = _csv_from_list(_skills_list(evt))

                # No capacity/languages fields in Event yet: use placeholders
                cap_current = 0
                cap_total   = 0
                languages_csv = ""

                # Use display label for status if choices are defined
                if hasattr(rec, "get_status_display"):
                    status_display = rec.get_status_display()
                else:
                    status_display = rec.status or ""

                ws.append([
                    volunteer_name,
                    getattr(evt, "name", "") if evt else "",
                    getattr(evt, "description", "") if evt else "",
                    getattr(evt, "location", "") if evt else "",
                    skills_csv,
                    getattr(evt, "urgency", "") if evt else "",
                    evt.event_date.strftime("%Y-%m-%d") if (evt and evt.event_date) else "",
                    f"{cap_current} / {cap_total}",
                    languages_csv,
                    status_display,
                ])

            # rough autofit
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(c.value or "")) for c in ws[col_letter])
                ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 2))

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}_{stamp}.xlsx"'
            return resp

        except ImportError:
            # CSV fallback
            import csv
            from io import StringIO

            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(headers)

            for rec in qs:
                evt = rec.event
                user = rec.volunteer

                if hasattr(user, "profile") and getattr(user.profile, "full_name", "").strip():
                    volunteer_name = user.profile.full_name.strip()
                else:
                    volunteer_name = getattr(user, "email", str(user))

                skills_csv = _csv_from_list(_skills_list(evt))

                cap_current = 0
                cap_total   = 0
                languages_csv = ""

                if hasattr(rec, "get_status_display"):
                    status_display = rec.get_status_display()
                else:
                    status_display = rec.status or ""

                writer.writerow([
                    volunteer_name,
                    getattr(evt, "name", "") if evt else "",
                    getattr(evt, "description", "") if evt else "",
                    getattr(evt, "location", "") if evt else "",
                    skills_csv,
                    getattr(evt, "urgency", "") if evt else "",
                    evt.event_date.strftime("%Y-%m-%d") if (evt and evt.event_date) else "",
                    f"{cap_current} / {cap_total}",
                    languages_csv,
                    status_display,
                ])

            resp = HttpResponse(
                sio.getvalue(),
                content_type="text/csv",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}_{stamp}.csv"'
            return resp

    # --------------------------- PDF export branch ---------------------------
    if request.GET.get("export_pdf") == "1":
        headers = [
            "Volunteer", "Event Name", "Description", "Location",
            "Required Skills", "Urgency", "Event Date",
            "Capacity (current / total)", "Languages", "Status",
        ]
        filename = "volunteer_history"
        if volunteer:
            filename += f"_{volunteer}"
        if status:
            filename += f"_{status}"

        rows = []
        for rec in qs:
            evt = rec.event
            volunteer_email = getattr(rec.volunteer, "email", str(rec.volunteer))
            skills_csv = _csv_from_list(_skills_list(evt))

            # still placeholders for capacity/languages
            cap_current = 0
            cap_total   = 0
            languages_csv = ""

            rows.append([
                volunteer_email,
                getattr(evt, "name", "") if evt else "",
                getattr(evt, "description", "") if evt else "",
                getattr(evt, "location", "") if evt else "",
                skills_csv,
                getattr(evt, "urgency", "") if evt else "",
                evt.event_date.strftime("%Y-%m-%d") if (evt and evt.event_date) else "",
                f"{cap_current} / {cap_total}",
                languages_csv,
                rec.status,
            ])

        return _pdf_table_response(headers, rows, filename_base=filename)

    # --------------------------- filters for dropdowns ---------------------------
    volunteer_users = (
        User.objects
        .filter(assignments__isnull=False)
        .select_related("profile")  # so we can use profile.full_name
        .distinct()
    )

    volunteers = []
    for u in volunteer_users:
        # Prefer Profile.full_name, fallback to email
        full_name = ""
        if hasattr(u, "profile") and getattr(u.profile, "full_name", "").strip():
            full_name = u.profile.full_name.strip()
        else:
            full_name = u.email

        volunteers.append(
            SimpleNamespace(
                id=u.email,  # used in the filter value (?volunteer=...)
                full_name=full_name,  # label shown in dropdown
            )
        )

    # status list stays like before, based on Assignment.status
    statuses = list(
        Assignment.objects.order_by().values_list("status", flat=True).distinct()
    )
    statuses = sorted([s for s in statuses if s]) or [
        Assignment.ASSIGNED,
        Assignment.ATTENDED,
        Assignment.NO_SHOW,
        Assignment.CANCELLED,
    ]

    # ---------------------------- build rows for HTML ----------------------------
    rows = []
    today = now().date()
    for r in qs:
        evt = r.event
        user = r.volunteer
        volunteer_email = getattr(user, "email", str(user))

        # Display name = Profile.full_name, fallback to email
        if hasattr(user, "profile") and getattr(user.profile, "full_name", "").strip():
            volunteer_display_name = user.profile.full_name.strip()
        else:
            volunteer_display_name = volunteer_email

        skills = _skills_list(evt)
        event_date = evt.event_date if (evt and evt.event_date) else None
        is_completed = bool(event_date and event_date < today)

        cap_current = 0
        cap_total = 0
        languages_list = []

        # Normalize status for display + styling
        raw_status = r.status or ""
        if hasattr(r, "get_status_display"):
            status_display = r.get_status_display() or raw_status
        else:
            status_display = raw_status

        s_lower = raw_status.replace("_", "-").lower()
        if s_lower in ("assigned", "registered"):
            status_class = "bg-secondary"
        elif s_lower == "attended":
            status_class = "text-bg-success"
        elif s_lower in ("no-show", "no show", "no_show"):
            status_class = "text-bg-danger"
        elif s_lower in ("cancelled", "canceled"):
            status_class = "text-bg-warning"
        else:
            status_class = "text-bg-dark"

        rows.append({
            "id": r.id,
            "volunteer_id": volunteer_email,  # filter identifier
            "volunteer_name": volunteer_display_name,  # Full Name in table
            "event_name": getattr(evt, "name", "") if evt else "",
            "description": getattr(evt, "description", "") if evt else "",
            "location": getattr(evt, "location", "") if evt else "",
            "required_skills": skills,
            "urgency": getattr(evt, "urgency", "") if evt else "",
            "event_date": event_date,
            "event_date_iso": event_date.isoformat() if event_date else "",
            "capacity": f"{cap_current} / {cap_total}",
            "languages": languages_list,
            "status": raw_status,
            "status_display": status_display,
            "status_class": status_class,
            "is_completed": is_completed,
        })


    return render(
        request,
        "volunteer_history.html",
        {
            "rows": rows,
            "volunteers": volunteers,
            "statuses": statuses,
            "count": len(rows),
            "filter": {
                "volunteer": volunteer,
                "status": status,
                "from_date": d_from.isoformat() if d_from else "",
            },
        },
    )

log = logging.getLogger(__name__)
@login_required
def profile_form(request):
    # Try to load existing profile for this user
    log.info("Using DB: %s", settings.DATABASES["default"]["NAME"])
    try:
        profile = getattr(request.user, "profile", None)
    except Profile.DoesNotExist:
        profile = None

    if request.method == "POST":
        form = ProfileForm(request.POST, instance=profile)

        posted_availability = request.POST.getlist("availability[]")  # e.g. ["2025-11-02", ...]
        posted_skills = request.POST.getlist("skills")                # matches <select multiple name="skills">

        if form.is_valid():
            prof = form.save(commit=False)
            prof.user = request.user
            prof.skills = posted_skills
            prof.availability = posted_availability
            prof.save()
            messages.success(request, "Profile saved to the database.")
            return redirect("profile_form")
        else:
            # show what failed in server logs
            log.warning("ProfileForm errors: %s", form.errors)
            messages.error(request, "Please fix the errors below.")
        selected_state = request.POST.get("state") or ""
        selected_skills = set(posted_skills)
        existing_availability = posted_availability
    else:
        # GET: prefill
        form = ProfileForm(instance=profile)
        selected_state = form.initial.get("state", "") if form.initial else ""
        selected_skills = set(profile.skills if profile else [])
        existing_availability = profile.availability if profile else []

    ctx = {
        "form": form,
        "STATE_CHOICES": STATE_CHOICES,
        "SKILL_CHOICES": SKILL_CHOICES,
        "selected_state": selected_state,
        "selected_skills": selected_skills,
        "existing_availability": existing_availability,
    }
    return render(request, "profile_form.html", ctx)

@login_required
@staff_member_required
def event_form(request):
    # --- CSV EXPORT BRANCH (GET ?export=1) ---
    if request.GET.get("export") == "1":
        import csv
        from io import StringIO

        # All events, ordered by date then name
        qs = Event.objects.all().order_by("event_date", "name")

        headers = [
            "Event Name",
            "Description",
            "Location",
            "Required Skills",
            "Urgency",
            "Event Date",
        ]

        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)

        for evt in qs:
            skills_csv = _csv_from_list(_skills_list(evt))

            writer.writerow([
                getattr(evt, "name", "") or "",
                getattr(evt, "description", "") or "",
                getattr(evt, "location", "") or "",
                skills_csv,
                getattr(evt, "urgency", "") or "",
                evt.event_date.strftime("%Y-%m-%d") if getattr(evt, "event_date", None) else "",
            ])

        stamp = now().strftime("%Y%m%d-%H%M%S")
        resp = HttpResponse(
            sio.getvalue(),
            content_type="text/csv",
        )
        resp["Content-Disposition"] = f'attachment; filename="events_{stamp}.csv"'
        return resp

        # --- PDF EXPORT BRANCH (GET ?export_pdf=1) ---
    if request.GET.get("export_pdf") == "1":
        # All events, ordered by date then name
        qs = Event.objects.all().order_by("event_date", "name")

        headers = [
            "Event Name",
            "Description",
            "Location",
            "Required Skills",
            "Urgency",
            "Event Date",
        ]

        rows = []
        for evt in qs:
            skills_csv = _csv_from_list(_skills_list(evt))
            rows.append([
                evt.name,
                evt.description,
                evt.location,
                skills_csv,
                evt.urgency,
                evt.event_date.strftime("%Y-%m-%d") if evt.event_date else "",
            ])

        return _pdf_table_response(headers, rows, filename_base="events")

    # --- NORMAL FORM HANDLING (EXISTING BEHAVIOR) ---
    form = EventForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        event = form.save()
        # If you need JSON for AJAX, add a branch here.
        return redirect("home")
    events = Event.objects.all().order_by("-event_date")

    return render(request, "event_form.html", {
        "form": form,
        "events": events})